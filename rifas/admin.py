from django.contrib import admin, messages
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as DjangoUserAdmin
from django import forms
from django.core.exceptions import ValidationError
from django.http import HttpResponse, HttpResponseRedirect
from django.utils.html import format_html
from django.utils import timezone
from django.utils.timezone import make_aware
from django.urls import reverse
from django.conf import settings
import secrets

from .models import AuditEvent, BankAccount, Customer, Raffle, RaffleCalculation, RaffleImage, RaffleOffer, SiteContent, Ticket, TicketPurchase, UserSecurity
from django.db import models
from django.db.models import OuterRef, Subquery
from datetime import datetime, time

# Admin UI (Spanish)
admin.site.site_header = "GanaHoyRD — Administración"
admin.site.site_title = "GanaHoyRD — Admin"
admin.site.index_title = "Panel de administración"
# Keep admin organized on desktop; CSS hides sidebar on mobile.
admin.site.enable_nav_sidebar = True


def _clean_invalid_raffle_id_filter(request) -> HttpResponseRedirect | None:
    """
    Django admin will show: 'Rifa con el ID “X” no existe' when a changelist receives an
    invalid related filter value like raffle__id__exact=X.
    This happens often due to preserved filters or copied URLs.
    If the ID does not exist, remove it and reload.
    """
    if request.method != "GET":
        return None
    try:
        raw = (request.GET.get("raffle__id__exact") or "").strip()
        if not raw:
            return None
        rid = int(raw)
        if rid <= 0:
            return None
    except Exception:
        return None
    try:
        if Raffle.objects.filter(id=rid).exists():
            return None
    except Exception:
        return None
    try:
        params = request.GET.copy()
        params.pop("raffle__id__exact", None)
        # Drop admin error flag param "e" if it was added.
        params.pop("e", None)
        return HttpResponseRedirect(f"{request.path}?{params.urlencode()}" if params else request.path)
    except Exception:
        return None


class RaffleOfferInline(admin.TabularInline):
    model = RaffleOffer
    extra = 0
    # El mínimo de compra se configura en la Rifa (separado de la oferta).
    fields = ("is_active", "buy_quantity", "bonus_quantity", "starts_at", "ends_at")


class RaffleImageInline(admin.TabularInline):
    model = RaffleImage
    extra = 0
    fields = ("image", "sort_order")
    ordering = ("sort_order", "created_at")


class RaffleAdminForm(forms.ModelForm):
    """
    Admin UX: allow picking only a date (no required hour).
    We store a safe default time internally (end of day).
    """

    draw_date = forms.DateField(
        label="Fecha de sorteo",
        required=True,
        widget=forms.DateInput(attrs={"type": "date"}, format="%Y-%m-%d"),
        help_text="Selecciona solo la fecha. La hora se guarda automáticamente.",
    )
    finished_at = forms.DateField(
        label="Fecha de finalización",
        required=False,
        widget=forms.DateInput(attrs={"type": "date"}, format="%Y-%m-%d"),
        help_text="Opcional. Selecciona solo la fecha. La hora se guarda automáticamente.",
    )

    class Meta:
        model = Raffle
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # When editing an existing raffle, prefill date inputs from the stored datetimes.
        # Otherwise, HTML5 date input can render blank and force re-entry.
        if getattr(self, "is_bound", False):
            return
        inst = getattr(self, "instance", None)
        if not inst or not getattr(inst, "pk", None):
            return
        try:
            dd = getattr(inst, "draw_date", None)
            if dd:
                try:
                    self.initial.setdefault("draw_date", timezone.localtime(dd).date())
                except Exception:
                    self.initial.setdefault("draw_date", dd.date())
        except Exception:
            pass
        try:
            fa = getattr(inst, "finished_at", None)
            if fa:
                try:
                    self.initial.setdefault("finished_at", timezone.localtime(fa).date())
                except Exception:
                    self.initial.setdefault("finished_at", fa.date())
        except Exception:
            pass

    def _date_to_dt_end_of_day(self, d):
        if not d:
            return None
        tz = timezone.get_current_timezone()
        dt = datetime.combine(d, time(23, 59, 0))
        try:
            return make_aware(dt, timezone=tz)
        except Exception:
            # If already aware or fails, fall back.
            return dt

    def clean_draw_date(self):
        d = self.cleaned_data.get("draw_date")
        return self._date_to_dt_end_of_day(d)

    def clean_finished_at(self):
        d = self.cleaned_data.get("finished_at")
        return self._date_to_dt_end_of_day(d) if d else None


class RaffleShowAllFilter(admin.SimpleListFilter):
    """
    Accept `show_all` param on Raffle changelist so admin doesn't redirect to `?e=1`.
    UI checkbox is injected via change_list_template (same pattern as TicketPurchase).
    """

    title = "Ver"
    parameter_name = "show_all"

    def lookups(self, request, model_admin):
        return [("1", "Todas (incluye rifas inactivas)")]

    def queryset(self, request, queryset):
        return queryset


@admin.register(Raffle)
class RaffleAdmin(admin.ModelAdmin):
    form = RaffleAdminForm
    list_display = (
        "title",
        "draw_date",
        "show_draw_date",
        "price_per_ticket",
        "ticket_counter",
        "is_active",
        "show_in_history",
        "show_in_my_tickets_search",
        "use_manual_progress",
        "manual_progress_percent",
        "created_at",
    )
    list_filter = (
        RaffleShowAllFilter,
        "is_active",
        "show_in_history",
        "show_in_my_tickets_search",
        "use_manual_progress",
        "show_draw_date",
    )
    search_fields = ("title", "slug")
    prepopulated_fields = {"slug": ("title",)}
    inlines = [RaffleImageInline, RaffleOfferInline]
    actions = ["show_in_history_action", "hide_from_history_action"]
    change_list_template = "admin/rifas/raffle/change_list.html"

    def changelist_view(self, request, extra_context=None):
        """
        Be defensive with preserved/foreign filters in querystring.
        If the user navigates here from another changelist, Django may carry
        `_changelist_filters` or unrelated lookups that can trigger
        "Rifa con el ID ... no existe" warnings. Strip those for this page.
        """
        if request.method == "GET":
            try:
                params = request.GET.copy()
                changed = False

                # Drop preserved filters from other changelists (not applicable here).
                if "_changelist_filters" in params:
                    params.pop("_changelist_filters", None)
                    changed = True

                # Drop any stray lookups referencing a `raffle` FK (this model has no such field).
                for k in list(params.keys()):
                    if k.startswith("raffle__"):
                        params.pop(k, None)
                        changed = True

                if changed:
                    return HttpResponseRedirect(f"{request.path}?{params.urlencode()}" if params else request.path)
            except Exception:
                pass
        return super().changelist_view(request, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Faster than counting Ticket rows: sum total_tickets of APPROVED purchases.
        qs = qs.annotate(
            sold_tickets_annot=models.Sum(
                "purchases__total_tickets",
                filter=models.Q(purchases__status=TicketPurchase.Status.APPROVED),
            )
        ).prefetch_related("offers")

        show_all = (request.GET.get("show_all") or "").strip() == "1"
        if show_all:
            return qs

        # Default: show only active raffles unless user explicitly filters is_active.
        if "is_active__exact" not in request.GET:
            return qs.filter(is_active=True)
        return qs

    @admin.action(description="Mostrar en historial")
    def show_in_history_action(self, request, queryset):
        queryset.update(show_in_history=True)

    @admin.action(description="Ocultar del historial")
    def hide_from_history_action(self, request, queryset):
        queryset.update(show_in_history=False)

    def save_model(self, request, obj, form, change):
        from .emails import send_winner_notification_sync
        from .audit import log_event

        prev_winner = None
        prev_active = None
        prev_hist = None
        if change and obj.pk:
            prev_winner, prev_active, prev_hist = (
                Raffle.objects.filter(pk=obj.pk)
                .values_list("winner_ticket_number", "is_active", "show_in_history")
                .first()
                or (None, None, None)
            )

        # Best-effort: validate raffle video duration <= 20s using metadata.
        # If it fails to read duration, we allow upload but recommend using MP4/WebM.
        if getattr(obj, "video", None):
            try:
                from mutagen import File as MutagenFile  # type: ignore

                f = obj.video.file
                try:
                    f.seek(0)
                except Exception:
                    pass
                meta = MutagenFile(f)
                length = float(getattr(getattr(meta, "info", None), "length", 0) or 0)
                if length and length > 20.0:
                    raise ValidationError("El video debe durar máximo 20 segundos.")
            except ValidationError:
                raise
            except Exception:
                # Don't block admin save for metadata issues.
                pass

            # No transcoding: you will upload MP4/MOV already compatible.
        res = super().save_model(request, obj, form, change)

        # Notify winner when winner ticket changes/gets set.
        try:
            new_winner = getattr(obj, "winner_ticket_number", None)
            became_inactive = (prev_active is True) and (obj.is_active is False)
            became_history = (prev_hist is False) and (obj.show_in_history is True)
            winner_changed = bool(new_winner and int(new_winner) != int(prev_winner or 0))

            # Your rule: send when raffle is inactive + shown in history + winner is set.
            should_notify = bool(new_winner and (obj.is_active is False) and bool(obj.show_in_history) and (winner_changed or became_inactive or became_history))

            if should_notify:
                t = (
                    Ticket.objects.select_related("purchase")
                    .filter(raffle=obj, number=int(new_winner))
                    .order_by("-id")
                    .first()
                )
                purchase = getattr(t, "purchase", None) if t else None
                if not t:
                    self.message_user(
                        request,
                        "No se pudo enviar: ese boleto no existe en la rifa. "
                        "Asegúrate de que la compra esté APROBADA (los boletos se crean al aprobar).",
                        level=messages.WARNING,
                    )
                    log_event(
                        request=request,
                        action=AuditEvent.Action.WINNER_SET,
                        raffle=obj,
                        from_status="",
                        to_status="",
                        notes="Intento de notificación: ticket no encontrado.",
                        extra={"winner_ticket_number": int(new_winner)},
                    )
                elif purchase and (getattr(purchase, "email", "") or "").strip():
                    inferred = ""
                    try:
                        inferred = request.build_absolute_uri("/").rstrip("/")
                    except Exception:
                        inferred = ""
                    ok, err = send_winner_notification_sync(
                        raffle=obj,
                        purchase=purchase,
                        ticket_display=obj.winner_ticket_display,
                        site_url=(getattr(settings, "SITE_URL", "") or inferred),
                    )
                    log_event(
                        request=request,
                        action=AuditEvent.Action.WINNER_SET,
                        raffle=obj,
                        purchase=purchase,
                        ticket=t,
                        notes=("Correo ganador enviado." if ok else f"Fallo correo ganador: {err}"),
                        extra={"winner_ticket_number": int(new_winner), "email": getattr(purchase, "email", ""), "sent": ok},
                    )
                    if ok:
                        self.message_user(request, "Correo de felicitación enviado al ganador.", level=messages.SUCCESS)
                    else:
                        self.message_user(request, f"No se pudo enviar correo al ganador: {err}", level=messages.ERROR)
                else:
                    self.message_user(
                        request,
                        "No se pudo enviar: el ganador no tiene email registrado en la compra.",
                        level=messages.WARNING,
                    )
        except Exception:
            # Never break admin save due to email issues.
            pass

        return res

    @admin.display(description="Boletos (vendidos/total)")
    def ticket_counter(self, obj: Raffle):
        if not obj.max_tickets:
            return "—"
        return f"{obj.sold_tickets}/{obj.max_tickets} ({obj.sold_percent}%)"

    def save_formset(self, request, form, formset, change):
        """
        Enforce max 3 images total per raffle:
        - optional cover image (Raffle.image) counts as 1 if set
        - plus inline gallery images (RaffleImage)
        """
        if formset.model is RaffleImage:
            raffle: Raffle = form.instance
            cover_count = 1 if getattr(raffle, "image", None) else 0

            submitted = 0
            for f in formset.forms:
                if not hasattr(f, "cleaned_data"):
                    continue
                if f.cleaned_data.get("DELETE"):
                    continue
                img = f.cleaned_data.get("image") or getattr(f.instance, "image", None)
                if img:
                    submitted += 1

            if cover_count + submitted > 3:
                raise ValidationError(
                    "Máximo 3 fotos por artículo (incluye la imagen principal). "
                    f"Ahora mismo: principal={cover_count}, galería={submitted}."
                )

        return super().save_formset(request, form, formset, change)


@admin.action(description="Aprobar compras seleccionadas")
def approve_purchases(modeladmin, request, queryset):
    from .emails import send_customer_purchase_status
    from .audit import log_event

    for purchase in queryset.select_related("raffle"):
        prev_status = purchase.status
        try:
            purchase.approve()
            log_event(
                request=request,
                action=AuditEvent.Action.PURCHASE_APPROVED,
                raffle=purchase.raffle,
                purchase=purchase,
                from_status=prev_status,
                to_status=purchase.status,
            )
            try:
                send_customer_purchase_status(purchase=purchase)
            except Exception:
                pass
        except ValueError as e:
            purchase.reject(notes=str(e))
            log_event(
                request=request,
                action=AuditEvent.Action.PURCHASE_REJECTED,
                raffle=purchase.raffle,
                purchase=purchase,
                from_status=prev_status,
                to_status=purchase.status,
                notes=str(e),
            )
            try:
                send_customer_purchase_status(purchase=purchase)
            except Exception:
                pass
            modeladmin.message_user(
                request,
                f"Compra #{purchase.id} rechazada: {e}",
                level=messages.WARNING,
            )


@admin.action(description="Rechazar compras seleccionadas")
def reject_purchases(modeladmin, request, queryset):
    from .emails import send_customer_purchase_status
    from .audit import log_event

    for purchase in queryset.select_related("raffle"):
        prev_status = purchase.status
        purchase.reject()
        log_event(
            request=request,
            action=AuditEvent.Action.PURCHASE_REJECTED,
            raffle=purchase.raffle,
            purchase=purchase,
            from_status=prev_status,
            to_status=purchase.status,
        )
        try:
            send_customer_purchase_status(purchase=purchase)
        except Exception:
            pass


@admin.action(description="Exportar compras a Excel (.xlsx)")
def export_ticket_purchases_xlsx(modeladmin, request, queryset):
    """
    Export selected TicketPurchase rows to Excel.
    """
    try:
        import openpyxl  # type: ignore
    except Exception:
        modeladmin.message_user(
            request,
            "Falta la dependencia openpyxl para exportar a Excel.",
            level=messages.ERROR,
        )
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"
    ws.append(["Nombre", "Teléfono", "Cantidad", "Monto (RD$)", "Banco", "Cuenta"])

    for p in queryset.select_related("bank_account").order_by("-created_at", "-id").iterator(chunk_size=1000):
        bank = getattr(p, "bank_account", None)
        ws.append(
            [
                getattr(p, "full_name", "") or "",
                getattr(p, "phone", "") or "",
                int(getattr(p, "quantity", 0) or 0),
                int(getattr(p, "total_amount", 0) or 0),
                (getattr(bank, "bank_name", "") or "") if bank else "",
                (getattr(bank, "account_number", "") or "") if bank else "",
            ]
        )

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="compras_boletos.xlsx"'
    wb.save(resp)
    return resp


class PhonePrefixFilter(admin.SimpleListFilter):
    title = "Prefijo"
    parameter_name = "phone_prefix"

    def lookups(self, request, model_admin):
        return [("809", "809"), ("829", "829"), ("849", "849")]

    def queryset(self, request, queryset):
        val = self.value()
        if not val:
            return queryset
        return queryset.filter(phone__startswith=val)


class TicketPurchaseShowAllFilter(admin.SimpleListFilter):
    """
    We use this filter only so Django admin accepts the `show_all` GET param.
    Otherwise, admin treats it as an unknown lookup and redirects to `?e=1`.
    """

    title = "Ver"
    parameter_name = "show_all"

    def lookups(self, request, model_admin):
        return [("1", "Todas (incluye rifas inactivas)")]

    def queryset(self, request, queryset):
        # Actual behavior is enforced in ModelAdmin.get_queryset.
        return queryset


class TicketShowAllFilter(admin.SimpleListFilter):
    """
    Accept `show_all` param on Ticket changelist.
    """

    title = "Ver"
    parameter_name = "show_all"

    def lookups(self, request, model_admin):
        return [("1", "Todos (incluye rifas inactivas)")]

    def queryset(self, request, queryset):
        return queryset


class AuditEventShowAllFilter(admin.SimpleListFilter):
    """
    Accept `show_all` param on AuditEvent changelist.
    """

    title = "Ver"
    parameter_name = "show_all"

    def lookups(self, request, model_admin):
        return [("1", "Todos (incluye rifas inactivas)")]

    def queryset(self, request, queryset):
        return queryset


@admin.register(TicketPurchase)
class TicketPurchaseAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "proof_link",
        "status",
        "id",
        "raffle",
        "full_name",
        "phone",
        "bank_account",
        "quantity",
        "promo_preview",
        "bonus_quantity",
        "total_tickets",
        "total_amount",
    )
    list_filter = (TicketPurchaseShowAllFilter, "status", "raffle", "bank_account", PhonePrefixFilter)
    search_fields = ("full_name", "phone", "email", "raffle__title", "bank_account__bank_name", "bank_account__account_number")
    search_help_text = "Busca por teléfono, nombre, rifa o banco."
    list_select_related = ("raffle", "bank_account")
    change_list_template = "admin/rifas/ticketpurchase/change_list.html"
    readonly_fields = (
        "created_at",
        "decided_at",
        "total_amount",
        "public_reference",
        "bonus_quantity",
        "total_tickets",
        "client_ip",
        "user_agent",
        "proof_preview",
    )
    actions = [approve_purchases, reject_purchases, export_ticket_purchases_xlsx]

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        show_all = (request.GET.get("show_all") or "").strip() == "1"
        if show_all:
            return qs
        # Default: show purchases for the active raffle only (fast FK filter).
        try:
            active = (
                Raffle.objects.filter(is_active=True)
                .order_by("-created_at")
                .only("id")
                .first()
            )
            if active and getattr(active, "id", None):
                return qs.filter(raffle_id=active.id)
        except Exception:
            pass
        return qs.none()

    def changelist_view(self, request, extra_context=None):
        """
        Default admin view: show pending purchases first.
        If the user already selected a status filter (or any status__exact), do not override.
        """
        cleaned = _clean_invalid_raffle_id_filter(request)
        if cleaned is not None:
            return cleaned

        # Export current filtered changelist to Excel (no selection needed)
        if request.method == "GET" and (request.GET.get("export") or "").strip() == "1":
            try:
                import openpyxl  # type: ignore
            except Exception:
                self.message_user(request, "Falta la dependencia openpyxl para exportar a Excel.", level=messages.ERROR)
            else:
                cl = self.get_changelist_instance(request)
                qs = cl.get_queryset(request).select_related("bank_account")

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Compras"
                ws.append(["Nombre", "Teléfono", "Cantidad", "Monto (RD$)", "Banco", "Cuenta"])

                for p in qs.order_by("-created_at", "-id").iterator(chunk_size=1000):
                    bank = getattr(p, "bank_account", None)
                    ws.append(
                        [
                            getattr(p, "full_name", "") or "",
                            getattr(p, "phone", "") or "",
                            int(getattr(p, "quantity", 0) or 0),
                            int(getattr(p, "total_amount", 0) or 0),
                            (getattr(bank, "bank_name", "") or "") if bank else "",
                            (getattr(bank, "account_number", "") or "") if bank else "",
                        ]
                    )

                resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                resp["Content-Disposition"] = 'attachment; filename="compras_boletos_filtradas.xlsx"'
                wb.save(resp)
                return resp

        if request.method == "GET" and "status__exact" not in request.GET:
            params = request.GET.copy()
            params["status__exact"] = TicketPurchase.Status.PENDING
            return HttpResponseRedirect(f"{request.path}?{params.urlencode()}")
        return super().changelist_view(request, extra_context=extra_context)

    @admin.display(description="Promoción (vista previa)")
    def promo_preview(self, obj: TicketPurchase):
        offer = obj.raffle.get_active_offer(obj.quantity) if obj.raffle_id else None
        if not offer:
            return "—"
        est = offer.bonus_for(obj.quantity) if obj.quantity else 0
        if est:
            return f"{offer.buy_quantity}+{offer.bonus_quantity} (gratis estimado: {est})"
        return f"{offer.buy_quantity}+{offer.bonus_quantity}"

    @admin.display(description="Comprobante")
    def proof_link(self, obj: TicketPurchase):
        f = getattr(obj, "proof_image", None)
        if not f or not getattr(f, "name", ""):
            return "—"
        try:
            # If the DB points to a missing file (common on Railway without a Volume),
            # don't show a broken link.
            if not f.storage.exists(f.name):
                return "No disponible"
            url = f.url
        except Exception:
            return "No disponible"
        return format_html('<a href="{}" target="_blank" rel="noopener">Ver</a>', url)

    @admin.display(description="Vista previa del comprobante")
    def proof_preview(self, obj: TicketPurchase):
        f = getattr(obj, "proof_image", None)
        if not f or not getattr(f, "name", ""):
            return "—"
        try:
            if not f.storage.exists(f.name):
                return "No disponible"
            url = f.url
        except Exception:
            return "No disponible"
        return format_html(
            '<a href="{0}" target="_blank" rel="noopener">'
            '<img src="{0}" alt="comprobante" style="max-width:360px; width:100%; border-radius:12px; border:1px solid rgba(255,255,255,.15);" />'
            "</a>",
            url,
        )

    def save_model(self, request, obj, form, change):
        from .emails import send_customer_purchase_status
        from .audit import log_event

        prev_status = None
        if change and obj.pk:
            prev_status = TicketPurchase.objects.filter(pk=obj.pk).values_list("status", flat=True).first()
        super().save_model(request, obj, form, change)
        if obj.status == TicketPurchase.Status.APPROVED and prev_status != TicketPurchase.Status.APPROVED:
            try:
                obj.apply_offer()
                obj.save(update_fields=["bonus_quantity", "total_tickets"])
                obj.generate_tickets_if_needed()
                try:
                    send_customer_purchase_status(purchase=obj)
                except Exception:
                    pass
            except ValueError as e:
                obj.reject(notes=str(e))
                self.message_user(request, f"No se pudo aprobar: {e}", level=messages.ERROR)
                try:
                    send_customer_purchase_status(purchase=obj)
                except Exception:
                    pass
        elif obj.status == TicketPurchase.Status.REJECTED and prev_status != TicketPurchase.Status.REJECTED:
            try:
                send_customer_purchase_status(purchase=obj)
            except Exception:
                pass

        # Audit status changes via manual edit.
        try:
            if change and prev_status and obj.status != prev_status:
                action = (
                    AuditEvent.Action.PURCHASE_APPROVED
                    if obj.status == TicketPurchase.Status.APPROVED
                    else AuditEvent.Action.PURCHASE_REJECTED
                    if obj.status == TicketPurchase.Status.REJECTED
                    else ""
                )
                if action:
                    log_event(
                        request=request,
                        action=action,
                        raffle=obj.raffle,
                        purchase=obj,
                        from_status=prev_status,
                        to_status=obj.status,
                        notes=(obj.admin_notes or ""),
                    )
        except Exception:
            pass

    def delete_model(self, request, obj):
        from .audit import log_event

        try:
            log_event(
                request=request,
                action=AuditEvent.Action.PURCHASE_DELETED,
                raffle=obj.raffle,
                purchase=obj,
                from_status=getattr(obj, "status", "") or "",
                to_status="deleted",
                notes="Eliminado desde admin.",
            )
        except Exception:
            pass
        return super().delete_model(request, obj)

    def delete_queryset(self, request, queryset):
        from .audit import log_event

        try:
            for p in queryset.select_related("raffle"):
                log_event(
                    request=request,
                    action=AuditEvent.Action.PURCHASE_DELETED,
                    raffle=p.raffle,
                    purchase=p,
                    from_status=getattr(p, "status", "") or "",
                    to_status="deleted",
                    notes="Eliminado en lote desde admin.",
                )
        except Exception:
            pass
        return super().delete_queryset(request, queryset)


@admin.register(SiteContent)
class SiteContentAdmin(admin.ModelAdmin):
    list_display = ("updated_at",)
    fieldsets = (
        ("Portada", {"fields": ("about_title", "about_body")}),
        ("Políticas (interno)", {"fields": ("policy_title", "policy_body")}),
        ("Métodos de pago (interno)", {"fields": ("payment_title", "payment_body")}),
        ("Detalles de pago", {"fields": ("payment_holder_name", "payment_account_type", "payment_currency")}),
        ("Términos y condiciones", {"fields": ("terms_title", "terms_body")}),
        ("Identidad", {"fields": ("site_logo",)}),
        ("CEO / Contacto", {"fields": ("ceo_name", "ceo_phone", "ceo_instagram_url", "ceo_tiktok_url")}),
    )


@admin.register(BankAccount)
class BankAccountAdmin(admin.ModelAdmin):
    list_display = ("bank_name", "owner_name", "account_number", "is_active", "sort_order")
    list_filter = ("is_active",)
    search_fields = ("bank_name", "owner_name", "account_number")


@admin.action(description="Exportar a Excel (.xlsx)")
def export_customers_xlsx(modeladmin, request, queryset):
    try:
        import openpyxl  # type: ignore
    except Exception:
        modeladmin.message_user(
            request,
            "Falta la dependencia openpyxl para exportar a Excel.",
            level=messages.ERROR,
        )
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Export only customer profile data (no purchase metrics)
    headers = ["Nombre", "Teléfono", "Email"]
    ws.append(headers)

    for c in queryset.order_by("-last_purchase_at", "-updated_at").iterator(chunk_size=1000):
        ws.append(
            [
                c.full_name,
                c.phone,
                c.email,
            ]
        )

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="clientes_ganahoyrd.xlsx"'
    wb.save(resp)
    return resp


@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ("full_name", "phone", "email")
    search_fields = ("full_name", "phone", "email")
    actions = [export_customers_xlsx]
    ordering = ("-last_purchase_at", "-updated_at", "-id")
    readonly_fields = (
        "first_purchase_at",
        "last_purchase_at",
        "total_purchases",
        "total_paid_tickets",
        "total_bonus_tickets",
        "total_amount",
    )
    fields = (
        "full_name",
        "phone",
        "email",
        "first_purchase_at",
        "last_purchase_at",
        "total_purchases",
        "total_paid_tickets",
        "total_bonus_tickets",
        "total_amount",
    )

    def has_add_permission(self, request):
        # Customers are created/updated automatically from purchases.
        # Hide "Añadir cliente" to avoid confusion.
        return False

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        # Defensive: if legacy data ever created duplicate rows with the same phone,
        # show only the most recently updated one per phone.
        latest_id_sq = (
            Customer.objects.filter(phone=OuterRef("phone"))
            .order_by("-updated_at", "-id")
            .values("id")[:1]
        )
        return qs.annotate(latest_id_annot=Subquery(latest_id_sq)).filter(id=models.F("latest_id_annot"))


@admin.register(Ticket)
class TicketAdmin(admin.ModelAdmin):
    list_display = ("raffle", "number_display", "purchase", "approved_by", "created_at")
    change_list_template = "admin/rifas/ticket/change_list.html"
    list_filter = (TicketShowAllFilter, "raffle")
    search_fields = ("purchase__full_name", "purchase__phone", "purchase__email", "raffle__title")
    list_select_related = ("raffle", "purchase")
    readonly_fields = ("number_display", "created_at")
    fields = ("raffle", "purchase", "number", "number_display", "created_at")

    def changelist_view(self, request, extra_context=None):
        cleaned = _clean_invalid_raffle_id_filter(request)
        if cleaned is not None:
            return cleaned

        # Export purchases (deduped) from current filtered tickets.
        if request.method == "GET" and (request.GET.get("export") or "").strip() == "1":
            try:
                import openpyxl  # type: ignore
            except Exception:
                self.message_user(request, "Falta la dependencia openpyxl para exportar a Excel.", level=messages.ERROR)
            else:
                cl = self.get_changelist_instance(request)
                # MySQL can error on DISTINCT with an implicit ORDER BY.
                # Remove ordering before distinct and keep it as a subquery (no big in-memory list).
                tqs = cl.get_queryset(request).select_related("purchase", "purchase__bank_account")
                purchase_ids_sq = (
                    tqs.exclude(purchase_id__isnull=True)
                    .order_by()
                    .values_list("purchase_id", flat=True)
                    .distinct()
                )

                pqs = (
                    TicketPurchase.objects.filter(id__in=purchase_ids_sq)
                    .select_related("bank_account")
                    .order_by("-created_at", "-id")
                )

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Compras"
                ws.append(["Nombre", "Teléfono", "Cantidad", "Monto (RD$)", "Banco", "Cuenta"])

                for p in pqs.iterator(chunk_size=1000):
                    bank = getattr(p, "bank_account", None)
                    ws.append(
                        [
                            getattr(p, "full_name", "") or "",
                            getattr(p, "phone", "") or "",
                            int(getattr(p, "quantity", 0) or 0),
                            int(getattr(p, "total_amount", 0) or 0),
                            (getattr(bank, "bank_name", "") or "") if bank else "",
                            (getattr(bank, "account_number", "") or "") if bank else "",
                        ]
                    )

                resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                resp["Content-Disposition"] = 'attachment; filename="compras_desde_boletos.xlsx"'
                wb.save(resp)
                return resp

        return super().changelist_view(request, extra_context=extra_context)

    def get_queryset(self, request):
        qs = super().get_queryset(request)

        # Annotate who approved the purchase (latest approval audit event).
        approved_by_sq = Subquery(
            AuditEvent.objects.filter(
                purchase_id=OuterRef("purchase_id"),
                action=AuditEvent.Action.PURCHASE_APPROVED,
            )
            .order_by("-created_at")
            .values("actor__username")[:1]
        )
        qs = qs.annotate(approved_by_annot=approved_by_sq)

        show_all = (request.GET.get("show_all") or "").strip() == "1"
        if show_all:
            return qs
        # Default: show tickets for the active raffle only.
        try:
            active = (
                Raffle.objects.filter(is_active=True)
                .order_by("-created_at")
                .only("id")
                .first()
            )
            if active and getattr(active, "id", None):
                return qs.filter(raffle_id=active.id)
        except Exception:
            pass
        return qs.none()

    @admin.display(description="Boleto")
    def number_display(self, obj: Ticket):
        return obj.display_number

    @admin.display(description="Aprobado por")
    def approved_by(self, obj: Ticket):
        return (getattr(obj, "approved_by_annot", None) or "—")

    def get_search_results(self, request, queryset, search_term):
        qs, use_distinct = super().get_search_results(request, queryset, search_term)
        term = (search_term or "").strip()
        digits = "".join(ch for ch in term if ch.isdigit())
        # Search by ticket number (accept 001/0001/etc.)
        if digits:
            try:
                n = int(digits)
                qs = qs | queryset.filter(number=n)
            except Exception:
                pass
        # Search by phone digits (ignore separators)
        if len(digits) >= 7:
            qs = qs | queryset.filter(purchase__phone__icontains=digits)
        return qs, use_distinct

    def delete_model(self, request, obj):
        from .audit import log_event

        try:
            log_event(
                request=request,
                action=AuditEvent.Action.TICKET_DELETED,
                raffle=obj.raffle,
                purchase=getattr(obj, "purchase", None),
                ticket=obj,
                notes=f"Eliminado boleto #{obj.display_number}.",
            )
        except Exception:
            pass
        return super().delete_model(request, obj)

    def delete_queryset(self, request, queryset):
        from .audit import log_event

        try:
            for t in queryset.select_related("raffle", "purchase"):
                log_event(
                    request=request,
                    action=AuditEvent.Action.TICKET_DELETED,
                    raffle=t.raffle,
                    purchase=getattr(t, "purchase", None),
                    ticket=t,
                    notes=f"Eliminado en lote boleto #{t.display_number}.",
                )
        except Exception:
            pass
        return super().delete_queryset(request, queryset)


@admin.register(AuditEvent)
class AuditEventAdmin(admin.ModelAdmin):
    list_display = ("created_at", "action", "actor", "raffle", "purchase", "ticket", "from_status", "to_status", "ip")
    change_list_template = "admin/rifas/auditevent/change_list.html"
    list_filter = (AuditEventShowAllFilter, "action", "created_at", "actor", "raffle")
    search_fields = ("purchase__public_reference", "purchase__full_name", "purchase__phone", "purchase__email", "ip", "user_agent")
    readonly_fields = (
        "created_at",
        "action",
        "actor",
        "raffle",
        "purchase",
        "ticket",
        "from_status",
        "to_status",
        "notes",
        "extra",
        "ip",
        "user_agent",
    )
    fields = readonly_fields

    def changelist_view(self, request, extra_context=None):
        cleaned = _clean_invalid_raffle_id_filter(request)
        if cleaned is not None:
            return cleaned
        return super().changelist_view(request, extra_context=extra_context)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        # allow view
        return True

    def has_delete_permission(self, request, obj=None):
        return False

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        show_all = (request.GET.get("show_all") or "").strip() == "1"
        if show_all:
            return qs
        # Default: hide inactive raffles, but keep events without a raffle.
        try:
            return qs.filter(models.Q(raffle__isnull=True) | models.Q(raffle__is_active=True))
        except Exception:
            return qs


@admin.register(RaffleCalculation)
class RaffleCalculationAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "raffle",
        "created_by",
        "ticket_price",
        "total_cost",
        "paid_tickets_needed",
        "bonus_tickets",
        "total_issued",
        "expected_profit",
    )
    list_filter = ("created_at", "raffle", "created_by")
    search_fields = ("raffle__title", "created_by__username")
    readonly_fields = (
        "created_at",
        "created_by",
        "raffle",
        "ticket_price",
        "product_cost",
        "shipping_cost",
        "advertising_cost",
        "other_costs",
        "desired_margin_percent",
        "offer_buy_quantity",
        "offer_bonus_quantity",
        "offer_min_paid_quantity",
        "total_cost",
        "revenue_needed",
        "break_even_tickets",
        "paid_tickets_needed",
        "bonus_tickets",
        "total_issued",
        "expected_revenue",
        "expected_profit",
        "max_tickets",
    )
    fields = readonly_fields

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


class UserSecurityInline(admin.StackedInline):
    model = UserSecurity
    can_delete = False
    extra = 0
    fields = ("force_password_change", "forced_at")
    readonly_fields = ("forced_at",)
    verbose_name_plural = "Seguridad"


# Extend Django's User admin to include "force password change"
User = get_user_model()
try:
    admin.site.unregister(User)
except admin.sites.NotRegistered:
    pass


@admin.register(User)
class UserAdmin(DjangoUserAdmin):
    inlines = (UserSecurityInline,)

    class AutoPasswordUserCreationForm(forms.ModelForm):
        """
        Create user without asking for password (we auto-generate + email it).
        """

        class Meta:
            model = User
            fields = ("username", "email", "first_name", "last_name", "is_active", "is_staff", "is_superuser", "groups", "user_permissions")

        def clean_email(self):
            email = (self.cleaned_data.get("email") or "").strip()
            if not email:
                raise forms.ValidationError("El email es requerido para crear el usuario (se enviará la contraseña temporal).")
            return email

        def save(self, commit=True):
            obj = super().save(commit=False)
            # Prevent empty password being saved (we will set a real one in save_model).
            try:
                obj.set_unusable_password()
            except Exception:
                pass
            if commit:
                obj.save()
                self.save_m2m()
            return obj

    add_form = AutoPasswordUserCreationForm

    add_fieldsets = (
        (None, {"classes": ("wide",), "fields": ("username", "email", "first_name", "last_name")}),
        ("Permisos", {"fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions")}),
    )

    def save_model(self, request, obj, form, change):
        """
        On user creation: generate temp password, email it, force change at next admin login.
        """
        creating = not change or not getattr(obj, "pk", None)
        super().save_model(request, obj, form, change)

        if creating:
            email = (getattr(obj, "email", "") or "").strip()
            if not email:
                self.message_user(
                    request,
                    "No se pudo enviar: el usuario no tiene email. Edita el usuario y agrega un email válido.",
                    level=messages.WARNING,
                )
                return

            temp_pwd = ("ADMIN-" + secrets.token_urlsafe(10)).replace("-", "").replace("_", "")[:12]
            obj.set_password(temp_pwd)
            obj.save(update_fields=["password"])

            try:
                sec, _created = UserSecurity.objects.get_or_create(user=obj)
                sec.force_password_change = True
                sec.password_hash_at_force = obj.password
                sec.save()
            except Exception:
                pass

            try:
                from .emails import send_new_admin_user_credentials

                inferred = ""
                try:
                    inferred = request.build_absolute_uri("/").rstrip("/")
                except Exception:
                    inferred = ""
                ok, err = send_new_admin_user_credentials(
                    to_email=email,
                    username=getattr(obj, "username", "") or "",
                    temp_password=temp_pwd,
                    site_url=(getattr(settings, "SITE_URL", "") or inferred),
                )
                if ok:
                    self.message_user(request, "Usuario creado. Se envió una contraseña temporal por correo.", level=messages.SUCCESS)
                else:
                    self.message_user(request, f"Usuario creado, pero no se pudo enviar el correo: {err}", level=messages.WARNING)
            except Exception:
                self.message_user(request, "Usuario creado, pero no se pudo enviar el correo (revisa logs).", level=messages.WARNING)

